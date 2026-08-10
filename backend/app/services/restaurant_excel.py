from collections import defaultdict
from datetime import UTC, date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models import Household, Wine, WineSale

HEADER_FILL = PatternFill("solid", fgColor="5A1429")
SECTION_FILL = PatternFill("solid", fgColor="EDE8DC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(size=18, bold=True, color="5A1429")
THIN_BORDER = Border(bottom=Side(style="thin", color="D9D3C4"))
MONEY_FORMAT = "#,##0.00"
DATE_FORMAT = "DD/MM/YYYY"


def safe_text(value: object) -> str:
    text = str(value or "")
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def money(value: Decimal | int | float | None) -> float:
    return float(value or 0)


def style_table(
    sheet, headers: list[str], rows: list[list[object]], money_columns: set[int] | None = None
) -> None:
    money_columns = money_columns or set()
    sheet.append(headers)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(len(rows) + 1, 1)}"
    for row in rows:
        sheet.append(row)
    for row in sheet.iter_rows(min_row=2):
        for index, cell in enumerate(row, start=1):
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=index == len(headers))
            if index in money_columns and isinstance(cell.value, int | float):
                cell.number_format = MONEY_FORMAT
            if isinstance(cell.value, date):
                cell.number_format = DATE_FORMAT
    for index, header in enumerate(headers, start=1):
        values = [safe_text(header), *(safe_text(row[index - 1]) for row in rows[:250])]
        width = min(max(max(map(len, values)) + 2, 10), 42)
        sheet.column_dimensions[get_column_letter(index)].width = width


def build_restaurant_excel(
    household: Household,
    sales_rows: list[tuple[WineSale, Wine]],
    inventory: list[Wine],
    from_date: date,
    to_date: date,
    locale: str,
    low_stock_threshold: int = 2,
) -> BytesIO:
    it = locale == "it"
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Riepilogo" if it else "Summary"
    summary["A1"] = "Vinaris · Export ristorante" if it else "Vinaris · Restaurant export"
    summary["A1"].font = TITLE_FONT
    summary["A3"] = "Cantina" if it else "Cellar"
    summary["B3"] = safe_text(household.name)
    summary["A4"] = "Periodo" if it else "Period"
    summary["B4"] = from_date
    summary["C4"] = to_date
    summary["B4"].number_format = DATE_FORMAT
    summary["C4"].number_format = DATE_FORMAT
    summary["A5"] = "Generato il" if it else "Generated at"
    summary["B5"] = datetime.now(UTC).replace(tzinfo=None)
    summary["B5"].number_format = "DD/MM/YYYY HH:MM"
    summary["A7"] = "Risultati per valuta" if it else "Results by currency"
    summary["A7"].fill = SECTION_FILL
    summary["A7"].font = Font(bold=True)

    totals: dict[str, dict[str, Decimal | int]] = defaultdict(
        lambda: {"revenue": Decimal("0"), "cost": Decimal("0"), "bottles": 0}
    )
    for sale, _wine in sales_rows:
        currency = (sale.currency or "CHF").upper()
        totals[currency]["revenue"] += sale.unit_sale_price * sale.quantity
        totals[currency]["cost"] += sale.unit_purchase_cost * sale.quantity
        totals[currency]["bottles"] += sale.quantity
    summary_headers = [
        "Valuta" if it else "Currency",
        "Bottiglie vendute" if it else "Bottles sold",
        "Ricavi" if it else "Revenue",
        "Costo" if it else "Cost",
        "Margine lordo" if it else "Gross margin",
        "Margine %" if it else "Margin %",
    ]
    summary_rows: list[list[object]] = []
    for currency, values in sorted(totals.items()):
        revenue = Decimal(values["revenue"])
        cost = Decimal(values["cost"])
        margin = revenue - cost
        summary_rows.append(
            [
                currency,
                int(values["bottles"]),
                money(revenue),
                money(cost),
                money(margin),
                float((margin / revenue * 100).quantize(Decimal("0.01"))) if revenue else 0,
            ]
        )
    for column, value in enumerate(summary_headers, start=1):
        cell = summary.cell(row=8, column=column, value=value)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_index, row in enumerate(summary_rows, start=9):
        for column, value in enumerate(row, start=1):
            summary.cell(row=row_index, column=column, value=value)
        for column in (3, 4, 5):
            summary.cell(row=row_index, column=column).number_format = MONEY_FORMAT
        summary.cell(row=row_index, column=6).number_format = '0.00"%"'
    for column, width in enumerate((12, 20, 16, 16, 18, 14), start=1):
        summary.column_dimensions[get_column_letter(column)].width = width

    sales_sheet = workbook.create_sheet("Vendite" if it else "Sales")
    sale_headers = [
        "Data" if it else "Date",
        "Vino" if it else "Wine",
        "Produttore" if it else "Producer",
        "Annata" if it else "Vintage",
        "Tipologia" if it else "Type",
        "Quantità" if it else "Quantity",
        "Prezzo unitario" if it else "Unit sale price",
        "Costo unitario" if it else "Unit cost",
        "Ricavi" if it else "Revenue",
        "Costo" if it else "Cost",
        "Margine" if it else "Margin",
        "Valuta" if it else "Currency",
        "Nota" if it else "Note",
    ]
    sale_rows = []
    for sale, wine in sales_rows:
        revenue = sale.unit_sale_price * sale.quantity
        cost = sale.unit_purchase_cost * sale.quantity
        sale_rows.append(
            [
                sale.sold_at,
                safe_text(wine.name),
                safe_text(wine.producer),
                safe_text(wine.vintage),
                safe_text(wine.type),
                sale.quantity,
                money(sale.unit_sale_price),
                money(sale.unit_purchase_cost),
                money(revenue),
                money(cost),
                money(revenue - cost),
                (sale.currency or "CHF").upper(),
                safe_text(sale.note),
            ]
        )
    style_table(sales_sheet, sale_headers, sale_rows, {7, 8, 9, 10, 11})

    producer_totals: dict[tuple[str, str], dict[str, Decimal | int]] = defaultdict(
        lambda: {"revenue": Decimal("0"), "cost": Decimal("0"), "bottles": 0}
    )
    for sale, wine in sales_rows:
        producer = wine.producer.strip() or ("Non indicato" if it else "Not specified")
        currency = (sale.currency or "CHF").upper()
        values = producer_totals[(producer, currency)]
        values["revenue"] += sale.unit_sale_price * sale.quantity
        values["cost"] += sale.unit_purchase_cost * sale.quantity
        values["bottles"] += sale.quantity
    producer_rows: list[list[object]] = []
    for (producer, currency), values in sorted(
        producer_totals.items(),
        key=lambda item: (int(item[1]["bottles"]), Decimal(item[1]["revenue"])),
        reverse=True,
    ):
        revenue = Decimal(values["revenue"])
        cost = Decimal(values["cost"])
        margin = revenue - cost
        producer_rows.append(
            [
                safe_text(producer),
                currency,
                int(values["bottles"]),
                money(revenue),
                money(cost),
                money(margin),
                float((margin / revenue * 100).quantize(Decimal("0.01"))) if revenue else 0,
            ]
        )
    producer_sheet = workbook.create_sheet("Produttori" if it else "Producers")
    style_table(
        producer_sheet,
        [
            "Produttore" if it else "Producer",
            "Valuta" if it else "Currency",
            "Bottiglie vendute" if it else "Bottles sold",
            "Ricavi" if it else "Revenue",
            "Costo" if it else "Cost",
            "Margine lordo" if it else "Gross margin",
            "Margine %" if it else "Margin %",
        ],
        producer_rows,
        {4, 5, 6},
    )
    for row_index in range(2, len(producer_rows) + 2):
        producer_sheet.cell(row=row_index, column=7).number_format = '0.00"%"'

    inventory_headers = [
        "Vino" if it else "Wine",
        "Produttore" if it else "Producer",
        "Annata" if it else "Vintage",
        "Tipologia" if it else "Type",
        "Regione" if it else "Region",
        "Denominazione" if it else "Appellation",
        "Giacenza" if it else "Stock",
        "Costo unitario" if it else "Unit cost",
        "Prezzo di vendita" if it else "Sale price",
        "Valuta" if it else "Currency",
        "Inizio finestra" if it else "Window start",
        "Fine finestra" if it else "Window end",
    ]

    def inventory_row(wine: Wine) -> list[object]:
        return [
            safe_text(wine.name),
            safe_text(wine.producer),
            safe_text(wine.vintage),
            safe_text(wine.type),
            safe_text(wine.region),
            safe_text(wine.appellation),
            wine.quantity,
            money(wine.price),
            money(wine.sale_price) if wine.sale_price is not None else None,
            (wine.currency or "CHF").upper(),
            wine.drink_from,
            wine.drink_to,
        ]

    inventory_sheet = workbook.create_sheet("Inventario" if it else "Inventory")
    inventory_rows = [inventory_row(wine) for wine in inventory]
    style_table(inventory_sheet, inventory_headers, inventory_rows, {8, 9})

    reorder_sheet = workbook.create_sheet("Da riordinare" if it else "Reorder")
    low_stock = sorted(
        (wine for wine in inventory if 0 < wine.quantity <= low_stock_threshold),
        key=lambda wine: (wine.quantity, wine.name.lower(), wine.vintage),
    )
    style_table(
        reorder_sheet, inventory_headers, [inventory_row(wine) for wine in low_stock], {8, 9}
    )

    missing_price_sheet = workbook.create_sheet("Prezzi mancanti" if it else "Missing prices")
    missing_price = [wine for wine in inventory if wine.sale_price is None]
    style_table(
        missing_price_sheet,
        inventory_headers,
        [inventory_row(wine) for wine in missing_price],
        {8, 9},
    )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
